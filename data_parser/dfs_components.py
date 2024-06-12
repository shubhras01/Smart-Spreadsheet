import openpyxl
import pandas as pd


class ExcelIslandFinder:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
        self.ws = self.wb.active
        self.visited = {}
        self.island_id = 0
        self.coords = {}
        self.configure_pandas_display()

    def configure_pandas_display(self):
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', 1000)
        pd.set_option('display.colheader_justify', 'center')
        pd.set_option('display.precision', 2)

    def has_borders(self, cell):
        """Check if the cell has any borders."""
        borders = cell.border
        return (borders.top.style is not None or borders.bottom.style is not None or
                borders.left.style is not None or borders.right.style is not None)

    def dfs(self, cell):
        """Use DFS to mark all cells connected with borders."""
        stack = [cell]
        directions = [(0, 1), (1, 0), (0, -1), (-1, 0)]  # Right, Down, Left, Up
        while stack:
            current_cell = stack.pop()
            if current_cell.coordinate in self.visited:
                continue
            self.visited[current_cell.coordinate] = self.island_id
            self.coords[current_cell.coordinate] = current_cell
            row, col = current_cell.row, current_cell.column
            for dr, dc in directions:
                new_row, new_col = row + dr, col + dc
                try:
                    neighbor = self.ws.cell(row=new_row, column=new_col)
                    if self.has_borders(neighbor) and neighbor.coordinate not in self.visited:
                        stack.append(neighbor)
                except IndexError:
                    continue
                except ValueError as e:
                    print(f"error happened: {e}")
                    continue

    def create_dataframe_from_box(self, box):
        """Creates a pandas DataFrame from cells within a bounding box."""
        min_row, max_row, min_col, max_col = box
        data = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell = self.ws.cell(row=row, column=col)
                row_data.append(cell.value)
            data.append(row_data)
        df = pd.DataFrame(data)
        return df

    def find_islands(self):
        """Find all islands of cells with borders."""
        for row in self.ws.iter_rows():
            for cell in row:
                if self.has_borders(cell) and cell.coordinate not in self.visited:
                    self.island_id += 1
                    self.dfs(cell)

    def get_dataframe_list(self):
        self.find_islands()
        print(f"Total islands found: {self.island_id}")
        box_to_cells = {}
        for k in self.visited:
            box_id = self.visited[k]
            box_to_cells.setdefault(box_id, [])
            box_to_cells[box_id].append(self.coords[k])
        all_dfs = []
        for k in box_to_cells:
            sorted_cells = sorted(box_to_cells[k], key=lambda x: (x.row, x.column))
            min_row = min(cell.row for cell in sorted_cells)
            max_row = max(cell.row for cell in sorted_cells)
            min_col = min(cell.column for cell in sorted_cells)
            max_col = max(cell.column for cell in sorted_cells)
            box = (min_row, max_row, min_col, max_col)
            print(f"Box {k}: {box}")
            df = self.create_dataframe_from_box(box)
            all_dfs.append(df)
        return all_dfs


if __name__ == "__main__":
    finder = ExcelIslandFinder('/home/shubhra/work/Smart-Spreadsheet/tests/example_0.xlsx')
    finder.get_dataframe_list()